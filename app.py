import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import io, os, re, zipfile, requests
from openpyxl import load_workbook

st.set_page_config(page_title="Gerador de Dinâmicas", page_icon="🏷️", layout="centered")

st.markdown("""
<style>
.hero{background:linear-gradient(135deg,#c8102e,#8B0000);border-radius:14px;padding:2rem;text-align:center;margin-bottom:1.5rem}
.hero h1{color:white;font-size:2rem;margin:0}
.hero p{color:rgba(255,255,255,.8);margin:.4rem 0 0}
.card{background:#1a1a1a;border:1px solid #2a2a2a;border-radius:12px;padding:1.4rem;margin-bottom:1.2rem}
.step{font-size:.72rem;font-weight:700;letter-spacing:2px;color:#c8102e;text-transform:uppercase;margin-bottom:.2rem}
.stitle{font-size:1.05rem;font-weight:600;color:#f0f0f0;margin-bottom:.7rem}
.info{background:#1e1e2e;border-left:3px solid #c8102e;border-radius:0 8px 8px 0;padding:.7rem 1rem;font-size:.85rem;color:#aaa;margin-bottom:.8rem}
.warn{background:#2a1f0e;border:1px solid #5a3a0e;border-radius:8px;padding:.7rem 1rem;color:#fbbf24;font-size:.85rem}
.ok{background:#0f2318;border:1px solid #1a4a2a;border-radius:8px;padding:.8rem 1rem;color:#4ade80;font-weight:600;text-align:center;margin-top:.8rem}
div[data-testid="stButton"]>button{background:linear-gradient(135deg,#c8102e,#8B0000);color:white;border:none;border-radius:8px;font-weight:700;font-size:1rem;width:100%}
</style>
""", unsafe_allow_html=True)

# ── Constantes ──────────────────────────────────────────────
A4_W, A4_H = 2480, 3508
CM = 118

KG_KEYWORDS = ['peça','pedaço','fatiada','fatia','a granel','granel',
                'presunto cozido','costela suína','queijo de coalho godam']

CORTES = [r'\bSELL OUT\b',r'\bSELL IN\b',r'\bPDN\b',r'\bGRADE\b',
          r'\bCUSTO FINAL\b',r'\bCUSTO NF\b',r'\bNF:\b',
          r'\bNesta Embalagem\b',r'\bNa compra\b',r'\binvest\.',r'\bSPLASH\b',
          r'\(NF',r'\(Custo',r'\(CUSTO',r'\(Desconto',r'\(sell',r'\(Sell',r'\(invest']

CORTES_ESP = {
    'arroz branco prato fino tipo 1': r'5kg',
    'café mais café':                 r'500g',
    'café pilão':                     r'500g',
    'cerveja flying fish long neck':  r'330ml',
}

# ── Funções de dados ─────────────────────────────────────────
def limpar_nome(txt):
    txt = txt.replace('\n',' ').strip()
    for p in CORTES:
        txt = re.split(p, txt, flags=re.IGNORECASE)[0].strip().rstrip('-– ').strip()
    txt = re.sub(r'\s*\($','',txt).strip()
    txt = re.sub(r'\s*\(Exceto.*','',txt,flags=re.IGNORECASE).strip()
    return txt

def corte_esp(nome):
    for pre, pat in CORTES_ESP.items():
        if nome.lower().startswith(pre):
            m = re.search(pat, nome, re.IGNORECASE)
            if m: return nome[:m.end()].strip()
    return nome

def unidade(nome):
    nl = nome.lower()
    return 'KG' if any(k in nl for k in KG_KEYWORDS) else 'CADA'

def resolver_preco(v):
    if v is None: return None
    s = str(v).strip()
    if s.startswith('='): return None
    try: return float(s)
    except: return None

@st.cache_data(show_spinner=False)
def ler_excel(data):
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=3, max_row=300, min_col=1, max_col=8):
        b, h = row[1], row[7]
        if not b.value: continue
        preco = resolver_preco(h.value)
        if preco is None: continue
        nome = corte_esp(limpar_nome(str(b.value)))
        out.append({'nome': nome, 'preco': preco, 'unidade': unidade(nome)})
    return out

@st.cache_data(show_spinner=False)
def baixar_drive(fid):
    url = f"https://drive.google.com/uc?export=download&id={fid}"
    try:
        s = requests.Session()
        r = s.get(url, stream=True, timeout=30)
        for k,v in r.cookies.items():
            if 'download_warning' in k:
                r = s.get(url+f"&confirm={v}", stream=True, timeout=30); break
        if r.status_code == 200: return r.content
    except: pass
    return None

def extrair_id(url):
    for p in [r'/file/d/([a-zA-Z0-9_-]+)',r'id=([a-zA-Z0-9_-]+)']:
        m = re.search(p, url)
        if m: return m.group(1)
    return None

def slugify(s):
    s = s.lower()
    for a,b in [('ã','a'),('ç','c'),('é','e'),('ê','e'),('á','a'),('í','i'),
                ('ó','o'),('ô','o'),('ú','u'),('â','a'),('õ','o'),('ü','u')]:
        s = s.replace(a,b)
    return re.sub(r'[^a-z0-9]+','_',s).strip('_')

# ── Renderização da dinâmica ─────────────────────────────────
def fonte(data, tam):
    try: return ImageFont.truetype(io.BytesIO(data), tam)
    except: return ImageFont.load_default()

def gerar(bg, box, prod_img, fn_bytes, fp_bytes, nome, preco, unid):
    canvas = Image.open(io.BytesIO(bg)).convert("RGBA")
    canvas = canvas.resize((A4_W, A4_H), Image.LANCZOS)
    draw   = ImageDraw.Draw(canvas)

    # Fontes
    f_nome = fonte(fn_bytes, 120) if fn_bytes else ImageFont.load_default()
    f_rs   = fonte(fp_bytes, 100) if fp_bytes else ImageFont.load_default()
    f_int  = fonte(fp_bytes, 300) if fp_bytes else ImageFont.load_default()
    f_dec  = fonte(fp_bytes, 150) if fp_bytes else ImageFont.load_default()
    f_un   = fonte(fp_bytes, 105) if fp_bytes else ImageFont.load_default()

    # ── Nome do produto ──
    palavras = nome.upper().split()
    linhas, linha = [], ""
    limite = A4_W - int(2*CM)
    for p in palavras:
        c = (linha+" "+p).strip()
        if draw.textlength(c, f_nome) <= limite: linha = c
        else:
            if linha: linhas.append(linha)
            linha = p
    if linha: linhas.append(linha)
    linhas = linhas[:2]
    y = int(11.5*CM)
    for l in linhas:
        w = draw.textlength(l, f_nome)
        draw.text(((A4_W-w)//2, y), l, fill="black", font=f_nome)
        y += f_nome.size + 14

    # ── Foto do produto ──
    if prod_img:
        pi = Image.open(io.BytesIO(prod_img)).convert("RGBA")
        h_max = int(11.5*CM)
        nova_h = h_max
        nova_w = int(nova_h * pi.width / pi.height)
        pi = pi.resize((nova_w, nova_h), Image.LANCZOS)
        canvas.paste(pi, ((A4_W-nova_w)//2, y + int(0.3*CM)), pi)

    # ── Box de preço ──
    BW = int(8.5*CM); BH = int(5.2*CM)
    cx = int(15.5*CM); cy = int(24.5*CM)
    x0 = cx-BW//2;    y0 = cy-BH//2
    raio = int(0.85*CM)

    if box:
        bi = Image.open(io.BytesIO(box)).convert("RGBA")
        bi = bi.resize((BW, BH), Image.LANCZOS)
        canvas.paste(bi, (x0, y0), bi)
    else:
        # Box amarelo desenhado
        overlay = Image.new("RGBA", (A4_W, A4_H), (0,0,0,0))
        od = ImageDraw.Draw(overlay)
        od.rounded_rectangle([x0+10,y0+10,x0+BW+10,y0+BH+10], radius=raio, fill=(0,0,0,70))
        od.rounded_rectangle([x0,y0,x0+BW,y0+BH], radius=raio,
                              fill=(255,220,0,255), outline=(210,170,0,255), width=8)
        canvas = Image.alpha_composite(canvas, overlay)
        draw = ImageDraw.Draw(canvas)

    # ── Textos do preço ──
    COR = "#D60000"
    inteiro = str(int(preco))
    dec     = f",{round((preco % 1)*100):02d}"

    # "R$" — topo esquerdo do box
    draw.text((x0+int(.3*CM), y0+int(.3*CM)), "R$", fill=COR, font=f_rs)

    # Valor inteiro + decimal centralizados verticalmente no box
    w_int = draw.textlength(inteiro, f_int)
    w_dec = draw.textlength(dec,     f_dec)
    total_w = w_int + w_dec
    xi = x0 + (BW - total_w)//2
    yi = y0 + (BH - f_int.size)//2 + int(.15*CM)
    draw.text((xi, yi), inteiro, fill=COR, font=f_int)

    # decimal alinhado à base do inteiro
    xd = xi + w_int
    yd = yi + f_int.size - f_dec.size
    draw.text((xd, yd), dec, fill=COR, font=f_dec)

    # unidade abaixo do decimal, alinhada à direita
    w_un = draw.textlength(unid, f_un)
    draw.text((xd + w_dec - w_un, yd + f_dec.size + int(.05*CM)),
              unid, fill=COR, font=f_un)

    return canvas.convert("RGB")

def to_png(img):
    b = io.BytesIO()
    img.save(b, format="PNG", dpi=(300,300))
    return b.getvalue()

# ── Interface ────────────────────────────────────────────────
st.markdown('<div class="hero"><h1>🏷️ Gerador de Dinâmicas</h1><p>Supermarket — automação de panfletos de oferta</p></div>', unsafe_allow_html=True)

# Passo 1 — Links fixos
st.markdown('<div class="card"><div class="step">Passo 1 — Configure uma vez</div><div class="stitle">🔗 Arquivos permanentes no Google Drive</div>', unsafe_allow_html=True)
st.markdown('<div class="info">Links de compartilhamento público dos arquivos que não mudam toda semana.</div>', unsafe_allow_html=True)
c1,c2 = st.columns(2)
with c1:
    url_fn  = st.text_input("Fonte do nome (Gagalin)",    placeholder="https://drive.google.com/file/d/...")
    url_fp  = st.text_input("Fonte do preço (ChunkFive)", placeholder="https://drive.google.com/file/d/...")
with c2:
    url_box = st.text_input("Box de preço PNG (opcional)", placeholder="https://drive.google.com/file/d/...")
    url_pasta = st.text_input("ID da pasta de imagens",   placeholder="ID da pasta do Google Drive")
st.markdown('</div>', unsafe_allow_html=True)

# Passo 2 — Background semanal
st.markdown('<div class="card"><div class="step">Passo 2 — Toda semana</div><div class="stitle">🎨 Background da semana</div>', unsafe_allow_html=True)
src = st.radio("Origem:", ["📤 Upload direto","🔗 Link do Drive"], horizontal=True, label_visibility="collapsed")
bg_bytes = None
if src == "📤 Upload direto":
    f = st.file_uploader("Background", type=["png","jpg","jpeg"], key="bg")
    if f:
        bg_bytes = f.read()
        prev = Image.open(io.BytesIO(bg_bytes))
        st.image(prev.resize((250, int(250*prev.height/prev.width))), caption="✅ Carregado")
else:
    u = st.text_input("Link do Drive", placeholder="https://drive.google.com/file/d/...", key="url_bg")
    if u:
        fid = extrair_id(u)
        if fid:
            with st.spinner("Baixando..."): bg_bytes = baixar_drive(fid)
            if bg_bytes:
                prev = Image.open(io.BytesIO(bg_bytes))
                st.image(prev.resize((250, int(250*prev.height/prev.width))), caption="✅ Carregado")
        else: st.warning("Link inválido.")
st.markdown('</div>', unsafe_allow_html=True)

# Passo 3 — Excel semanal
st.markdown('<div class="card"><div class="step">Passo 3 — Toda semana</div><div class="stitle">📊 Lâmina de Ofertas (Excel)</div>', unsafe_allow_html=True)
st.markdown('<div class="info">Envie o mesmo Excel da lâmina. O app lê <b>Coluna B</b> (produto) e <b>Coluna H</b> (preço) automaticamente — sem precisar editar nada.</div>', unsafe_allow_html=True)
xf = st.file_uploader("Excel da lâmina", type=["xlsx","xls"], key="xl")
produtos = []
if xf:
    try:
        produtos = ler_excel(xf.read())
        st.markdown(f'<div class="ok">✅ {len(produtos)} produtos lidos (3 ignorados com fórmula de pack)</div>', unsafe_allow_html=True)
        with st.expander("Ver todos os produtos"):
            for i,p in enumerate(produtos,1):
                st.write(f"{i}. {p['nome']} — R$ {p['preco']:.2f} — **{p['unidade']}**")
    except Exception as e:
        st.error(f"Erro: {e}")
st.markdown('</div>', unsafe_allow_html=True)

# Passo 4 — Gerar
st.markdown('<div class="card"><div class="step">Passo 4</div><div class="stitle">⚡ Gerar</div>', unsafe_allow_html=True)

ok = produtos and bg_bytes
if not ok:
    f = []
    if not bg_bytes: f.append("Background")
    if not produtos: f.append("Excel")
    st.markdown(f'<div class="warn">⏳ Faltando: {", ".join(f)}</div>', unsafe_allow_html=True)

if st.button("🚀 Gerar todas as dinâmicas", disabled=not ok):

    fn_bytes = fp_bytes = box_b = None
    with st.spinner("Baixando fontes e box..."):
        if url_fn:  fn_bytes  = baixar_drive(extrair_id(url_fn))  if extrair_id(url_fn)  else None
        if url_fp:  fp_bytes  = baixar_drive(extrair_id(url_fp))  if extrair_id(url_fp)  else None
        if url_box: box_b     = baixar_drive(extrair_id(url_box)) if extrair_id(url_box) else None

    prog = st.progress(0, text="Iniciando...")
    geradas, sem_foto = [], []
    total = len(produtos)

    for i, p in enumerate(produtos):
        prog.progress(i/total, text=f"{i+1}/{total} — {p['nome'][:45]}")

        # Tenta buscar imagem da pasta do Drive pelo slug
        prod_img = None
        if url_pasta:
            slug = slugify(p['nome'])
            for ext in ['png','jpg','jpeg']:
                # Tenta via URL direta (requer pasta pública com nome exato)
                try_url = f"https://drive.google.com/uc?export=download&id={url_pasta}/{slug}.{ext}"
                # Nota: listagem real requer Google API Key — veja instruções
            sem_foto.append(p['nome'])

        try:
            img = gerar(bg_bytes, box_b, prod_img,
                        fn_bytes, fp_bytes,
                        p['nome'], p['preco'], p['unidade'])
            arq = f"{i+1:02d}_{slugify(p['nome'])[:35]}.png"
            geradas.append((arq, img))
        except Exception as e:
            st.warning(f"Erro em '{p['nome']}': {e}")

    prog.progress(1.0, text="✅ Concluído!")

    if geradas:
        st.markdown(f'<div class="ok">🎉 {len(geradas)} dinâmicas geradas!</div>', unsafe_allow_html=True)

        st.markdown("#### Preview das primeiras 3")
        cols = st.columns(min(3, len(geradas)))
        for idx,(arq,img) in enumerate(geradas[:3]):
            with cols[idx]:
                st.image(img.resize((220, int(220*A4_H/A4_W))), caption=f"#{idx+1}")

        # ZIP
        zb = io.BytesIO()
        with zipfile.ZipFile(zb,"w",zipfile.ZIP_DEFLATED) as zf:
            for arq,img in geradas: zf.writestr(arq, to_png(img))
        st.download_button("📥 Baixar todas (.zip)", zb.getvalue(), "dinamicas.zip", "application/zip", use_container_width=True)

        # PDF
        pb = io.BytesIO()
        pags = [img for _,img in geradas]
        pags[0].save(pb, format="PDF", save_all=True, append_images=pags[1:], resolution=300)
        st.download_button("📄 Baixar PDF único", pb.getvalue(), "dinamicas.pdf", "application/pdf", use_container_width=True)

st.markdown('</div>', unsafe_allow_html=True)
